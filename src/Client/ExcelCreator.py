import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

def export_to_excel(
    patients_db, analyses_db, half_year_db, output_file,
    patient_fields, half_year_fields
):
    # --- 1. Загружаем пациентов ---
    conn_pat = sqlite3.connect(patients_db)
    cur_pat = conn_pat.cursor()
    cur_pat.execute("SELECT * FROM PATIENTS")
    patients = cur_pat.fetchall()
    patient_columns = [desc[0] for desc in cur_pat.description]
    conn_pat.close()

    # --- 2. Загружаем анализы ---
    conn_an = sqlite3.connect(analyses_db)
    cur_an = conn_an.cursor()
    cur_an.execute("SELECT * FROM ANALYSES")
    analyses = cur_an.fetchall()
    conn_an.close()

    # --- 3. Загружаем полугодовые ---
    conn_half = sqlite3.connect(half_year_db)
    cur_half = conn_half.cursor()
    cur_half.execute("SELECT * FROM HALF_YEAR_ANALYSES")
    half_year = cur_half.fetchall()
    half_year_columns = [desc[0] for desc in cur_half.description]
    conn_half.close()

    # --- 4. Сопоставляем по patient_id ---
    analyses_dict = {}
    for row in analyses:
        pid = row[1]
        analyses_dict.setdefault(pid, []).append(row)

    half_year_dict = {}
    for row in half_year:
        pid = row[1]
        half_year_dict.setdefault(pid, []).append(row)

    # --- 5. Формируем строки ---
    all_rows = []
    for patient in patients:
        pid = patient[0]

        patient_data = {col: val for col, val in zip(patient_columns, patient)}
        patient_data.pop("id", None)

        patient_analyses = analyses_dict.get(pid, [])
        analyses_text = "; ".join([f"{r[2]}={r[3]} ({r[4]})" for r in patient_analyses]) if patient_analyses else ""

        patient_half = half_year_dict.get(pid, [])

        if patient_half:
            for i, half in enumerate(patient_half):
                half_data = {col: val for col, val in zip(half_year_columns, half)}
                half_data.pop("id", None)
                half_data.pop("patient_id", None)

                row = (
                        [patient_data.get(key, "") if i == 0 else "" for key, _ in patient_fields] +
                        [""] +
                        [half_data.get(key, "") for key, _ in half_year_fields] +
                        [""] +
                        [analyses_text if i == 0 else ""]
                )
                all_rows.append(row)
        else:
            empty_half = [""] * len(half_year_fields)
            row = (
                    [patient_data.get(key, "") for key, _ in patient_fields] +
                    [""] +
                    empty_half +
                    [""] +
                    [analyses_text]
            )
            all_rows.append(row)

    # --- 6. Заголовки ---
    headers = (
        [label for _, label in patient_fields]
        + [" "]
        + [label for _, label in half_year_fields]
        + [" "]
        + ["Анализы"]
    )

    # --- 7. Создаём DataFrame ---
    df = pd.DataFrame(all_rows, columns=headers)
    df.to_excel(output_file, index=False)

    # --- 8. Форматирование Excel ---
    wb = load_workbook(output_file)
    ws = wb.active

    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = Alignment(wrap_text=True)  # Перенос текста
        if col_name == " ":
            for row_idx in range(1, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).fill = gray_fill  # Серый фон

    wb.save(output_file)
    print(f"✅ Файл сохранён и отформатирован: {output_file}")

def export_to_excel_horizontal(
    patients_db, analyses_db, half_year_db, output_file,
    patient_fields, half_year_fields
):
    # --- 1. Загружаем пациентов ---
    conn_pat = sqlite3.connect(patients_db)
    cur_pat = conn_pat.cursor()
    cur_pat.execute("SELECT * FROM PATIENTS")
    patients = cur_pat.fetchall()
    patient_columns = [desc[0] for desc in cur_pat.description]
    conn_pat.close()

    # --- 2. Загружаем анализы ---
    conn_an = sqlite3.connect(analyses_db)
    cur_an = conn_an.cursor()
    cur_an.execute("SELECT * FROM ANALYSES")
    analyses = cur_an.fetchall()
    conn_an.close()

    # --- 3. Загружаем полугодовые ---
    conn_half = sqlite3.connect(half_year_db)
    cur_half = conn_half.cursor()
    cur_half.execute("SELECT * FROM HALF_YEAR_ANALYSES")
    half_year = cur_half.fetchall()
    half_year_columns = [desc[0] for desc in cur_half.description]
    conn_half.close()

    # --- 4. Сопоставляем по patient_id ---
    analyses_dict = {}
    for row in analyses:
        pid = row[1]
        analyses_dict.setdefault(pid, []).append(row)

    half_year_dict = {}
    for row in half_year:
        pid = row[1]
        half_year_dict.setdefault(pid, []).append(row)

    # --- 5. Формируем строки ---
    all_rows = []
    max_half_surveys = max((len(v) for v in half_year_dict.values()), default=0)

    for patient in patients:
        pid = patient[0]
        patient_data = {col: val for col, val in zip(patient_columns, patient)}
        patient_data.pop("id", None)
        patient_analyses = analyses_dict.get(pid, [])
        analyses_text = "; ".join([f"{r[2]}={r[3]} ({r[4]})" for r in patient_analyses]) if patient_analyses else ""
        patient_half = half_year_dict.get(pid, [])

        row = [patient_data.get(key, "") for key, _ in patient_fields]

        for i in range(max_half_surveys):
            row.append("")  # Серый столбец
            if i < len(patient_half):
                half = patient_half[i]
                half_data = {col: val for col, val in zip(half_year_columns, half)}
                half_data.pop("id", None)
                half_data.pop("patient_id", None)
                row.extend([half_data.get(key, "") for key, _ in half_year_fields])
            else:
                row.extend([""] * len(half_year_fields))

        row.append("")  # Серый столбец перед анализами
        row.append(analyses_text)
        all_rows.append(row)

    # --- 6. Заголовки ---
    headers = [label for _, label in patient_fields]
    for i in range(max_half_surveys):
        headers.append("")  # Серый столбик
        headers.extend([label for _, label in half_year_fields])
    headers.append("")  # Серый столбец перед анализами
    headers.append("Анализы")

    # --- 7. Создаём DataFrame ---
    df = pd.DataFrame(all_rows, columns=headers)
    df.to_excel(output_file, index=False)

    # --- 8. Форматирование Excel ---
    wb = load_workbook(output_file)
    ws = wb.active
    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    ALL_CELL_NAMES = ["До терапии"] + [f"{i} контроль" for i in range(1, max_half_surveys + 1)] + [""]
    cellNameIdx = 0
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = Alignment(wrap_text=True)
        if col_name == "":
            cell.value = ALL_CELL_NAMES[cellNameIdx]
            cellNameIdx += 1

            ws.cell(row=1, column=col_idx).fill = gray_fill

            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).fill = gray_fill

    wb.save(output_file)
    print(f"✅ Горизонтальный Excel сохранён: {output_file}")


def ExcelExport(output):
    from src.Client.botUTILS import BOT
    bot = BOT()
    export_to_excel_horizontal(bot.Patients.DB_Path, bot.Analyses.DB_Path, bot.HalfYearAnalyses.DB_Path,
                             output, bot.Patients.FIELDS, bot.HalfYearAnalyses.FIELDS)